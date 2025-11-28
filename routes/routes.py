# routes/routes.py
import time
import json
from datetime import datetime
from typing import Optional, Tuple
from flask import Blueprint, request, jsonify
from Postgress.Tables import (
    TenantCredentials,
    TenantUsers,
    StorageTargets,
    Templates,
    OperationLogs,
    OperationType,
    RenderStatus,
)
from Auth.Microsoft_Graph_Auth import MicrosoftGraphAuthenticator
from Services.graph_services import GraphServices, GraphAPIError
from Services.excel_section_writer import procesar_excel_completo
from validators.payload import (
    ALLOWED_IDENTIFIER_RE,
    CELL_REF_RE,
    MAX_CLIENT_FIELD_LENGTH,
    MAX_TENANT_NAME_LENGTH,
    MAX_DEST_FILE_NAME_LENGTH,
    MAX_TARGET_ALIAS_LENGTH,
    MAX_TEMPLATE_KEY_LENGTH,
    build_dest_file_name,
    new_correlation_id,
    require_fields,
    validate_cell_map_or_raise,
    validate_data_dict,
    validate_location_selector,
    validate_naming_dict,
    validate_section_data,
    validate_string_field,
)

graph_bp = Blueprint("graph", __name__)

# ------------------ Helpers ------------------ #
def _join_storage_path(*segments: str) -> str:
    cleaned = []
    for segment in segments:
        if segment is None:
            continue
        stripped = str(segment).strip()
        if not stripped:
            continue
        cleaned.append(stripped.strip("/"))
    if not cleaned:
        return ""
    return "/".join(cleaned)


def _resolve_graph_target(storage) -> Tuple[Optional[str], Optional[str]]:
    """Devuelve (drive_id, target_user_id) soportando los campos nuevos y legacy."""
    if storage is None:
        return None, None

    location_type = getattr(storage, "location_type", None)
    location_identifier = getattr(storage, "location_identifier", None)
    if location_type in {"drive", "user"} and location_identifier:
        if location_type == "drive":
            return location_identifier, None
        return None, location_identifier

    return None, getattr(storage, "target_user_id", None)


# ------------------ ROUTES ------------------ #

@graph_bp.post("/excel/render-upload")
def render_upload():
    """
    1) Lee tenant/template/target/tenantuser de DB
    2) Descarga template desde OneDrive del cliente
    3) Rellena en memoria con 'data' (modo legacy) o 'sections' (modo dinámico)
    4) Sube el archivo final al destino (conflictBehavior del template)
    5) Inserta log en render_logs

    MODO LEGACY (referencias estáticas):
    curl -X POST http://localhost:8000/graph/excel/render-upload 
        -H "Content-Type: application/json" 
        -d '{
            "client_key": "eunoia",
            "template_key": "waman_prueba",
            "tenant_name": "Eunoia",
            "data": {
                "A1": "Contoso",
                "Hoja1!B3": 42,
                "Hoja1!C5": "=SUM(A1:A10)"
            }
        }'

    MODO SECCIONES (dinámico con marcadores):
    curl -X POST http://localhost:8000/graph/excel/render-upload 
        -H "Content-Type: application/json" 
        -d '{
            "client_key": "eunoia",
            "template_key": "invoice_template",
            "tenant_name": "Eunoia",
            "sections": {
                "header": {
                    "nombre": "ACME S.A.",
                    "rfc": "ACM123456",
                    "fecha": "2025-11-22"
                },
                "items": [
                    {"concepto": "Servicio A", "monto": 1000, "iva": "=B5*0.16"},
                    {"concepto": "Servicio B", "monto": 1500, "iva": "=B6*0.16"}
                ],
                "footer": {
                    "subtotal": 2500,
                    "total": "=B10+SUM(C5:C6)"
                }
            },
            "naming": {"periodo": "2024-05"},
            "target_alias": "finance",
            "requested_by": "api-client-123"
        }'
    """
    body = request.get_json() or {}
    
    # Detectar si usa modo secciones o modo legacy
    uses_sections = "sections" in body
    
    if uses_sections:
        # Modo secciones dinámicas
        err = require_fields(body, ["client_key", "template_key", "sections", "tenant_name"])
    else:
        # Modo legacy (referencias estáticas)
        err = require_fields(body, ["client_key", "template_key", "data", "tenant_name"])
    
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("template_key", body.get("template_key"), max_length=MAX_TEMPLATE_KEY_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("tenant_name", body.get("tenant_name"), max_length=MAX_TENANT_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    # Validar según el modo
    if uses_sections:
        err = validate_section_data(body.get("sections"))
        if err:
            return jsonify({"error": err}), 400
    else:
        # Modo legacy
        err = validate_data_dict(body.get("data"))
        if err:
            return jsonify({"error": err}), 400

        try:
            validate_cell_map_or_raise(body["data"])
        except ValueError as ve:
            return jsonify({"error": str(ve)}), 400

    naming_provided = body.get("naming") not in (None, {})
    naming_err, sanitized_naming = validate_naming_dict(body.get("naming"))
    if naming_err:
        return jsonify({"error": naming_err}), 400
    body["naming"] = sanitized_naming

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400
        alias_clean = target_alias.strip()
        if alias_clean and not ALLOWED_IDENTIFIER_RE.match(alias_clean):
            return jsonify({"error": "target_alias contiene caracteres no permitidos"}), 400
        body["target_alias"] = alias_clean or None

    loc_err, (normalized_type, ident_clean) = validate_location_selector(body.get("location_type"), body.get("location_identifier"))
    if loc_err:
        return jsonify({"error": loc_err}), 400
    body["location_type"] = normalized_type
    body["location_identifier"] = ident_clean

    corr_id = new_correlation_id()
    t0 = time.perf_counter()

    db = request.environ.get("db_session")

    dest_file_name = None
    template = None

    try:
        # 1) Config del tenant
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        target_alias = body.get("target_alias")
        location_type = body.get("location_type")
        location_identifier = body.get("location_identifier")
        tenant_user = None
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no está configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
            if not storage:
                return jsonify({"error": f"No hay destino configurado para el alias '{target_alias}'"}), 400
        else:
            if location_type and location_identifier:
                storage = (
                    storage_query
                    .filter_by(location_type=location_type, location_identifier=location_identifier)
                    .first()
                )
                if not storage:
                    return jsonify({"error": "No hay destino configurado para la ubicación indicada"}), 400
            else:
                storage = storage_query.first()

        template = (
            db.query(Templates)
            .filter_by(client_key=body["client_key"], template_key=body["template_key"])
            .first()
        )
        if not storage or not template:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        # 2) Token MSAL
        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        # 3) Descargar template
        full_template_path = _join_storage_path(
            template.template_folder_path,
            template.template_file_name,
        )
        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        tpl_bytes, ms_id_download = gs.download_file_bytes(
            full_template_path,
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        # 4) Render en memoria (modo secciones o legacy)
        if uses_sections:
            # Modo secciones dinámicas
            section_configs = template.cell_mapping.get("sections", {}) if template.cell_mapping else {}
            
            if not section_configs:
                return jsonify({
                    "error": f"Template '{template.template_key}' no tiene secciones configuradas en cell_mapping"
                }), 400
            
            try:
                # procesar_excel_completo returns a BytesIO
                output_io = procesar_excel_completo(template_bytes=tpl_bytes, secciones=body["sections"], configuracion=section_configs)
                filled_bytes = output_io.getvalue()
            except ValueError as ve:
                return jsonify({"error": f"Error al procesar secciones: {str(ve)}"}), 400
        else:
            # Modo legacy (referencias estáticas)
            filled_bytes = gs.render_in_memory(tpl_bytes, body["data"])

        # 5) Calcular nombre de archivo final según patrón configurable
        try:
            dest_file_name = build_dest_file_name(template, body, naming_provided=naming_provided)
        except ValueError as ve:
            return jsonify({"error": str(ve)}), 400

        dest_path = _join_storage_path(storage.default_dest_folder_path, dest_file_name)
        upload_result, ms_id_upload = gs.upload_file_bytes(
            filled_bytes,
            dest_path,
            conflict_behavior=template.default_conflict_behavior,
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        # 6) Guardar log
        duration_ms = int((time.perf_counter() - t0) * 1000)
        
        # Guardar data_json según el modo usado
        data_for_log = body.get("sections") if uses_sections else body.get("data")
        
        log_row = OperationLogs(
            client_key=body["client_key"],
            template_id=template.id,
            operation_type=OperationType.copy_template,
            input_data=data_for_log,
            output_data={
                "drive_item_id": upload_result.get("id"),
                "web_url": upload_result.get("webUrl"),
                "dest_file_name": dest_file_name,
            },
            status=RenderStatus.success,
            requested_by=body.get("requested_by", "eco-agent"),
            executed_at=datetime.utcnow(),
            duration_ms=duration_ms,
            ms_request_ids={"download_template": ms_id_download, "upload_file": ms_id_upload},
        )

        db.add(log_row)
        # commit/rollback lo maneja tu teardown; si prefieres commit explícito, descomenta:
        # db.commit()

        return jsonify(
            {
                "message": "OK",
                "driveItem": upload_result,
                "correlation_id": corr_id,
                "duration_ms": duration_ms,
                "ms_request_ids": {
                    "download_template": ms_id_download,
                    "upload_file": ms_id_upload,
                },
            }
        ), 200

    except GraphAPIError as ge:
        try:
            duration_ms = int((time.perf_counter() - t0) * 1000)
            log_row = OperationLogs(
                client_key=body.get("client_key"),
                template_id=template.id if template else None,
                operation_type=OperationType.copy_template,
                input_data=body.get("data", {}),
                status=RenderStatus.error,
                error_message=f"{ge.message} | ms-request-id={ge.ms_request_id}",
                requested_by=body.get("requested_by", "eco-agent"),
                executed_at=datetime.utcnow(),
                duration_ms=duration_ms,
                output_data={"dest_file_name": dest_file_name},
            )
            db.add(log_row)
        except Exception:
            pass

        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        try:
            duration_ms = int((time.perf_counter() - t0) * 1000)
            log_row = OperationLogs(
                client_key=body.get("client_key"),
                template_id=template.id if template else None,
                operation_type=OperationType.copy_template,
                input_data=body.get("data", {}),
                status=RenderStatus.error,
                error_message=str(e),
                requested_by=body.get("requested_by", "eco-agent"),
                executed_at=datetime.utcnow(),
                duration_ms=duration_ms,
                output_data={"dest_file_name": dest_file_name},
            )
            db.add(log_row)
        except Exception:
            pass

        return jsonify({"error": str(e), "correlation_id": corr_id}), 500


@graph_bp.post("/excel/write-cells")
def write_cells():
    """
    1) Lee tenant/target de DB
    2) Valida el mapa de celdas (A1 / Hoja!A1)
    3) Escribe directamente en el Excel existente vía Graph Excel API
    4) Inserta log en render_logs (sin template_id)

    curl -X POST http://localhost:8000/graph/excel/write-cells \
        -H "Content-Type: application/json" \
        -H "X-Api-Key: opcional-si-quieres-saltar-rate-limit" \
        -d '{
            "client_key": "contoso",
            "dest_file_name": "reporte-abril.xlsx",
            "data": {
            "A1": "Contoso S.A.",
            "Hoja1!B3": 12850,
            "Hoja1!C4": "OK"
            },
            "target_alias": "finance",                              opcional
            "location_type": "user",                                opcional
            "location_identifier": "finance-team@contoso.com",      opcional
            "requested_by": "api-client-123"                        opcional
    }'
    """
    body = request.get_json() or {}
    err = require_fields(body, ["client_key", "tenant_name", "dest_file_name", "data"])
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("dest_file_name", body.get("dest_file_name"), max_length=MAX_DEST_FILE_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_data_dict(body.get("data"))
    if err:
        return jsonify({"error": err}), 400

    # Validación de direcciones y valores
    try:
        validate_cell_map_or_raise(body["data"])
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400

    dest_file_name = body["dest_file_name"].strip()
    if any(sep in dest_file_name for sep in ("/", "\\")) or dest_file_name.startswith(".") or ".." in dest_file_name:
        return jsonify({"error": "dest_file_name inválido"}), 400
    if not dest_file_name.lower().endswith(".xlsx"):
        return jsonify({"error": "dest_file_name debe terminar en .xlsx"}), 400
    body["dest_file_name"] = dest_file_name

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400
        alias_clean = target_alias.strip()
        if alias_clean and not ALLOWED_IDENTIFIER_RE.match(alias_clean):
            return jsonify({"error": "target_alias contiene caracteres no permitidos"}), 400
        body["target_alias"] = alias_clean or None

    loc_err, (normalized_type, ident_clean) = validate_location_selector(
        body.get("location_type"), body.get("location_identifier")
    )
    if loc_err:
        return jsonify({"error": loc_err}), 400
    body["location_type"] = normalized_type
    body["location_identifier"] = ident_clean

    corr_id = new_correlation_id()
    t0 = time.perf_counter()

    db = request.environ.get("db_session")

    try:
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        target_alias = body.get("target_alias")
        location_type = body.get("location_type")
        location_identifier = body.get("location_identifier")
        tenant_user = None
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no está configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
            if not storage:
                return jsonify({"error": f"No hay destino configurado para el alias '{target_alias}'"}), 400
        else:
            if location_type and location_identifier:
                storage = (
                    storage_query
                    .filter_by(location_type=location_type, location_identifier=location_identifier)
                    .first()
                )
                if not storage:
                    return jsonify({"error": "No hay destino configurado para la ubicación indicada"}), 400
            else:
                storage = storage_query.first()

        if not storage:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        # Token MSAL
        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        # Ruta al archivo existente
        full_dest_path = _join_storage_path(
            storage.default_dest_folder_path,
            body["dest_file_name"],
        )

        # Escribir todas las celdas
        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        result, ms_request_ids = gs.write_cells_graph(
            full_dest_path=full_dest_path,
            data=body["data"],
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        written = result.get("written", {})
        error_cells = {cell: info for cell, info in written.items() if info.get("status") == "error"}

        log_status = RenderStatus.SUCCESS if not error_cells else RenderStatus.ERROR
        error_summary = None
        if error_cells:
            try:
                error_summary = json.dumps(error_cells)
            except Exception:
                error_summary = str(error_cells)
            if error_summary and len(error_summary) > 1000:
                error_summary = error_summary[:997] + "..."

        # Log (sin template_id porque es edición directa)
        duration_ms = int((time.perf_counter() - t0) * 1000)
        log_row = OperationLogs(
            client_key=body["client_key"],
            template_id=None,  # importante: None, no 0
            operation_type=OperationType.update_cell,
            input_data=body["data"],
            output_data={"dest_file_name": body["dest_file_name"]},
            status=RenderStatus.success if log_status == RenderStatus.SUCCESS else RenderStatus.error,
            requested_by=body.get("requested_by", "eco-agent"),
            executed_at=datetime.utcnow(),
            duration_ms=duration_ms,
            error_message=error_summary,
        )
        db.add(log_row)
        # db.commit()  # opcional; tu teardown ya hace commit si no hay excepción

        response_payload = {
            "message": "OK" if not error_cells else "Parcial",
            "written": written,
            "correlation_id": corr_id,
            "duration_ms": duration_ms,
            "ms_request_ids": ms_request_ids,
        }

        status_code = 200 if not error_cells else 207  # 207 Multi-Status para resultados mixtos
        return jsonify(response_payload), status_code

    except GraphAPIError as ge:
        try:
            duration_ms = int((time.perf_counter() - t0) * 1000)
            log_row = OperationLogs(
                client_key=body.get("client_key"),
                template_id=None,
                operation_type=OperationType.update_cell,
                input_data=body.get("data", {}),
                status=RenderStatus.error,
                error_message=f"{ge.message} | ms-request-id={ge.ms_request_id}",
                requested_by=body.get("requested_by", "eco-agent"),
                executed_at=datetime.utcnow(),
                duration_ms=duration_ms,
                output_data={"dest_file_name": body.get("dest_file_name")},
            )
            db.add(log_row)
        except Exception:
            pass

        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        # Intentar registrar el error también
        try:
            duration_ms = int((time.perf_counter() - t0) * 1000)
            log_row = OperationLogs(
                client_key=body.get("client_key"),
                template_id=None,
                operation_type=OperationType.update_cell,
                input_data=body.get("data", {}),
                status=RenderStatus.error,
                error_message=str(e),
                requested_by=body.get("requested_by", "eco-agent"),
                executed_at=datetime.utcnow(),
                duration_ms=duration_ms,
                output_data={"dest_file_name": body.get("dest_file_name")},
            )
            db.add(log_row)
            # db.commit()  # opcional, tu teardown gestionará rollback si vuelve a fallar
        except Exception:
            pass

        return jsonify({"error": str(e), "correlation_id": corr_id}), 500


@graph_bp.post("/excel/read-cells")
def read_cells():
    """
    Permite verificar que un archivo Excel en el destino contiene los valores esperados.
    El agente envía la ubicación del archivo y el json `cell_mapping_fill` con los valores
    esperados; la ruta compara contra el contenido actual del workbook.
    """
    body = request.get_json() or {}
    err = require_fields(body, ["client_key", "tenant_name", "dest_file_name", "cell_mapping_fill"])
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("tenant_name", body.get("tenant_name"), max_length=MAX_TENANT_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("dest_file_name", body.get("dest_file_name"), max_length=MAX_DEST_FILE_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    expected_cells = body.get("cell_mapping_fill")
    if not isinstance(expected_cells, dict):
        return jsonify({"error": "cell_mapping_fill debe ser un diccionario"}), 400

    err = validate_data_dict(expected_cells)
    if err:
        return jsonify({"error": err}), 400

    try:
        validate_cell_map_or_raise(expected_cells)
    except ValueError as ve:
        return jsonify({"error": str(ve)}), 400

    dest_file_name = body["dest_file_name"].strip()
    if any(sep in dest_file_name for sep in ("/", "\\")) or dest_file_name.startswith(".") or ".." in dest_file_name:
        return jsonify({"error": "dest_file_name inválido"}), 400
    if not dest_file_name.lower().endswith(".xlsx"):
        return jsonify({"error": "dest_file_name debe terminar en .xlsx"}), 400
    body["dest_file_name"] = dest_file_name

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400
        alias_clean = target_alias.strip()
        if alias_clean and not ALLOWED_IDENTIFIER_RE.match(alias_clean):
            return jsonify({"error": "target_alias contiene caracteres no permitidos"}), 400
        body["target_alias"] = alias_clean or None

    loc_err, (normalized_type, ident_clean) = validate_location_selector(
        body.get("location_type"), body.get("location_identifier")
    )
    if loc_err:
        return jsonify({"error": loc_err}), 400
    body["location_type"] = normalized_type
    body["location_identifier"] = ident_clean

    corr_id = new_correlation_id()
    t0 = time.perf_counter()

    db = request.environ.get("db_session")

    try:
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        target_alias = body.get("target_alias")
        location_type = body.get("location_type")
        location_identifier = body.get("location_identifier")
        tenant_user = None
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no está configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
            if not storage:
                return jsonify({"error": f"No hay destino configurado para el alias '{target_alias}'"}), 400
        else:
            if location_type and location_identifier:
                storage = (
                    storage_query
                    .filter_by(location_type=location_type, location_identifier=location_identifier)
                    .first()
                )
                if not storage:
                    return jsonify({"error": "No hay destino configurado para la ubicación indicada"}), 400
            else:
                storage = storage_query.first()

        if not storage:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        full_dest_path = _join_storage_path(
            storage.default_dest_folder_path,
            body["dest_file_name"],
        )

        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        read_result, ms_ids = gs.read_cells_graph(
            full_dest_path=full_dest_path,
            cells=list(expected_cells.keys()),
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        verification = {}
        mismatches = {}
        cells_info = read_result.get("cells", {})
        for cell, info in cells_info.items():
            expected_value = expected_cells.get(cell)
            cell_entry = {
                "status": info.get("status"),
                "expected": expected_value,
                "actual": info.get("value"),
            }
            if info.get("status") == "ok":
                matches = info.get("value") == expected_value
                cell_entry["matches"] = matches
                if not matches:
                    mismatches[cell] = {
                        "expected": expected_value,
                        "actual": info.get("value"),
                    }
            else:
                cell_entry["message"] = info.get("message")
                cell_entry["http_status"] = info.get("http_status")
                if info.get("ms_request_id"):
                    cell_entry["ms_request_id"] = info.get("ms_request_id")
                mismatches[cell] = {
                    "expected": expected_value,
                    "error": info.get("message"),
                    "http_status": info.get("http_status"),
                }
            verification[cell] = cell_entry

        duration_ms = int((time.perf_counter() - t0) * 1000)
        all_match = not mismatches

        response_payload = {
            "message": "OK" if all_match else "Differences detected",
            "matches_all": all_match,
            "verification": verification,
            "mismatches": mismatches,
            "correlation_id": corr_id,
            "duration_ms": duration_ms,
            "ms_request_ids": ms_ids,
        }

        return jsonify(response_payload), 200

    except GraphAPIError as ge:
        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        return jsonify({"error": str(e), "correlation_id": corr_id}), 500



@graph_bp.post("/excel/read-range")
def read_range():
    """
    Lee valores de celdas específicas sin verificar contra valores esperados.
    Envía una lista en "cells" y devuelve el valor/status de cada una.
    """
    body = request.get_json() or {}
    err = require_fields(body, ["client_key", "tenant_name", "dest_file_name", "cells"])
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("tenant_name", body.get("tenant_name"), max_length=MAX_TENANT_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("dest_file_name", body.get("dest_file_name"), max_length=MAX_DEST_FILE_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    cells = body.get("cells")
    if not isinstance(cells, list) or not cells:
        return jsonify({"error": "cells debe ser una lista no vacía de direcciones"}), 400
    for cell in cells:
        if not isinstance(cell, str) or not CELL_REF_RE.match(cell):
            return jsonify({"error": f"Dirección de celda inválida: '{cell}'. Usa 'A1' o 'Hoja!B2'."}), 400

    dest_file_name = body["dest_file_name"].strip()
    if any(sep in dest_file_name for sep in ("/", "\\")) or dest_file_name.startswith(".") or ".." in dest_file_name:
        return jsonify({"error": "dest_file_name inválido"}), 400
    if not dest_file_name.lower().endswith(".xlsx"):
        return jsonify({"error": "dest_file_name debe terminar en .xlsx"}), 400
    body["dest_file_name"] = dest_file_name

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400
        alias_clean = target_alias.strip()
        if alias_clean and not ALLOWED_IDENTIFIER_RE.match(alias_clean):
            return jsonify({"error": "target_alias contiene caracteres no permitidos"}), 400
        body["target_alias"] = alias_clean or None

    loc_err, (normalized_type, ident_clean) = validate_location_selector(
        body.get("location_type"), body.get("location_identifier")
    )
    if loc_err:
        return jsonify({"error": loc_err}), 400
    body["location_type"] = normalized_type
    body["location_identifier"] = ident_clean

    corr_id = new_correlation_id()
    t0 = time.perf_counter()

    db = request.environ.get("db_session")

    try:
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        target_alias = body.get("target_alias")
        location_type = body.get("location_type")
        location_identifier = body.get("location_identifier")
        tenant_user = None
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no está configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
            if not storage:
                return jsonify({"error": f"No hay destino configurado para el alias '{target_alias}'"}), 400
        else:
            if location_type and location_identifier:
                storage = (
                    storage_query
                    .filter_by(location_type=location_type, location_identifier=location_identifier)
                    .first()
                )
                if not storage:
                    return jsonify({"error": "No hay destino configurado para la ubicación indicada"}), 400
            else:
                storage = storage_query.first()

        if not storage:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        full_dest_path = _join_storage_path(
            storage.default_dest_folder_path,
            body["dest_file_name"],
        )

        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        read_result, ms_ids = gs.read_cells_graph(
            full_dest_path=full_dest_path,
            cells=cells,
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        cells_info = read_result.get("cells", {})
        error_cells = {cell: info for cell, info in cells_info.items() if info.get("status") != "ok"}

        duration_ms = int((time.perf_counter() - t0) * 1000)
        response_payload = {
            "message": "OK" if not error_cells else "Partial",
            "cells": cells_info,
            "error_cells": error_cells,
            "correlation_id": corr_id,
            "duration_ms": duration_ms,
            "ms_request_ids": ms_ids,
        }
        status_code = 200 if not error_cells else 207
        return jsonify(response_payload), status_code

    except GraphAPIError as ge:
        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        return jsonify({"error": str(e), "correlation_id": corr_id}), 500



@graph_bp.post("/excel/insert-rows")
def insert_rows():
    """
    Inserta una o varias filas desplazando hacia abajo a partir de start_cell (A2 o Hoja!A2)
    y escribe los valores provistos.
    """
    body = request.get_json() or {}
    err = require_fields(body, ["client_key", "tenant_name", "dest_file_name", "start_cell"])
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("tenant_name", body.get("tenant_name"), max_length=MAX_TENANT_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("dest_file_name", body.get("dest_file_name"), max_length=MAX_DEST_FILE_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    start_cell = body.get("start_cell")
    if not isinstance(start_cell, str) or not CELL_REF_RE.match(start_cell):
        return jsonify({"error": "start_cell inválido; usa 'A2' o 'Hoja!A2'"}), 400

    rows = body.get("rows")
    row_count = body.get("row_count")
    merge_ranges = body.get("merge_ranges")
    max_cols = 0
    if rows is None:
        if row_count is None:
            return jsonify({"error": "Debes enviar rows o row_count"}), 400
        if not isinstance(row_count, int) or row_count <= 0:
            return jsonify({"error": "row_count debe ser un entero > 0"}), 400
    else:
        if not isinstance(rows, list) or not rows:
            return jsonify({"error": "rows debe ser una lista de filas (listas) no vacía"}), 400
        for idx, row in enumerate(rows):
            if not isinstance(row, list):
                return jsonify({"error": f"rows[{idx}] debe ser una lista"}), 400
            if not row:
                return jsonify({"error": f"rows[{idx}] no puede estar vacío"}), 400
            for val in row:
                if not (isinstance(val, (str, int, float, bool)) or val is None):
                    return jsonify({"error": f"Tipo no soportado en rows[{idx}]: {type(val).__name__}"}), 400
            max_cols = max(max_cols, len(row))
        if max_cols == 0:
            return jsonify({"error": "rows debe contener al menos una columna"}), 400

    dest_file_name = body["dest_file_name"].strip()
    if any(sep in dest_file_name for sep in ("/", "\\")) or dest_file_name.startswith(".") or ".." in dest_file_name:
        return jsonify({"error": "dest_file_name inválido"}), 400
    if not dest_file_name.lower().endswith(".xlsx"):
        return jsonify({"error": "dest_file_name debe terminar en .xlsx"}), 400
    body["dest_file_name"] = dest_file_name

    if merge_ranges is not None:
        if not isinstance(merge_ranges, list) or not all(isinstance(r, str) for r in merge_ranges):
            return jsonify({"error": "merge_ranges debe ser una lista de strings"}), 400

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400
        alias_clean = target_alias.strip()
        if alias_clean and not ALLOWED_IDENTIFIER_RE.match(alias_clean):
            return jsonify({"error": "target_alias contiene caracteres no permitidos"}), 400
        body["target_alias"] = alias_clean or None

    loc_err, (normalized_type, ident_clean) = validate_location_selector(
        body.get("location_type"), body.get("location_identifier")
    )
    if loc_err:
        return jsonify({"error": loc_err}), 400
    body["location_type"] = normalized_type
    body["location_identifier"] = ident_clean

    corr_id = new_correlation_id()
    t0 = time.perf_counter()

    db = request.environ.get("db_session")

    try:
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        target_alias = body.get("target_alias")
        location_type = body.get("location_type")
        location_identifier = body.get("location_identifier")
        tenant_user = None
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no está configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
            if not storage:
                return jsonify({"error": f"No hay destino configurado para el alias '{target_alias}'"}), 400
        else:
            if location_type and location_identifier:
                storage = (
                    storage_query
                    .filter_by(location_type=location_type, location_identifier=location_identifier)
                    .first()
                )
                if not storage:
                    return jsonify({"error": "No hay destino configurado para la ubicación indicada"}), 400
            else:
                storage = storage_query.first()

        if not storage:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        full_dest_path = _join_storage_path(
            storage.default_dest_folder_path,
            body["dest_file_name"],
        )

        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        insert_result, ms_ids = gs.insert_rows_graph(
            full_dest_path=full_dest_path,
            start_cell=start_cell,
            rows=rows if rows is not None else None,
            row_count=row_count if rows is None else None,
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
            merge_ranges=merge_ranges,
        )

        duration_ms = int((time.perf_counter() - t0) * 1000)
        response_payload = {
            **insert_result,
            "correlation_id": corr_id,
            "duration_ms": duration_ms,
            "ms_request_ids": ms_ids,
        }
        return jsonify(response_payload), 200

    except GraphAPIError as ge:
        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        return jsonify({"error": str(e), "correlation_id": corr_id}), 500

@graph_bp.post("/excel/find-markers")
def find_markers():
    """
    Endpoint de utilidad para descubrir marcadores en un template Excel.
    Retorna lista de textos que parecen ser marcadores (terminan en ':' o contienen palabras clave).
    
    Útil para configurar cell_mapping.sections sin abrir el Excel manualmente.
    
    curl -X POST http://localhost:8000/graph/excel/find-markers 
        -H "Content-Type: application/json" 
        -d '{
            "client_key": "eunoia",
            "template_key": "invoice_template",
            "tenant_name": "Eunoia"
        }'
    """
    body = request.get_json() or {}
    err = require_fields(body, ["client_key", "template_key", "tenant_name"])
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("client_key", body.get("client_key"), max_length=MAX_CLIENT_FIELD_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("template_key", body.get("template_key"), max_length=MAX_TEMPLATE_KEY_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    err = validate_string_field("tenant_name", body.get("tenant_name"), max_length=MAX_TENANT_NAME_LENGTH)
    if err:
        return jsonify({"error": err}), 400

    target_alias = body.get("target_alias")
    if target_alias is not None:
        err = validate_string_field("target_alias", target_alias, max_length=MAX_TARGET_ALIAS_LENGTH)
        if err:
            return jsonify({"error": err}), 400

    loc_err, (normalized_type, ident_clean) = validate_location_selector(
        body.get("location_type"), body.get("location_identifier")
    )
    if loc_err:
        return jsonify({"error": loc_err}), 400

    corr_id = new_correlation_id()
    db = request.environ.get("db_session")

    try:
        creds = (
            db.query(TenantCredentials)
            .filter_by(client_key=body["client_key"], enabled=True)
            .first()
        )
        if not creds:
            return jsonify({"error": "Configuración incompleta en DB"}), 400

        template = (
            db.query(Templates)
            .filter_by(client_key=body["client_key"], template_key=body["template_key"])
            .first()
        )
        if not template:
            return jsonify({"error": "Template no encontrado"}), 400

        storage_query = (
            db.query(StorageTargets)
            .filter_by(client_key=body["client_key"], tenant_id=creds.id)
        )
        
        if target_alias:
            tenant_user = (
                db.query(TenantUsers)
                .filter_by(tenant_id=creds.id, alias=target_alias)
                .first()
            )
            if not tenant_user:
                return jsonify({"error": f"target_alias '{target_alias}' no configurado"}), 400
            storage = storage_query.filter_by(tenant_user_id=tenant_user.id).first()
        elif normalized_type and ident_clean:
            storage = (
                storage_query
                .filter_by(location_type=normalized_type, location_identifier=ident_clean)
                .first()
            )
        else:
            storage = storage_query.first()

        if not storage:
            return jsonify({"error": "No hay destino de almacenamiento configurado"}), 400

        auth = MicrosoftGraphAuthenticator(
            creds.tenant_id, creds.app_client_id, creds.app_client_secret
        )
        token = auth.get_access_token()
        gs = GraphServices(access_token=token, correlation_id=corr_id)

        full_template_path = _join_storage_path(
            template.template_folder_path,
            template.template_file_name,
        )
        drive_id, target_user_graph_id = _resolve_graph_target(storage)

        tpl_bytes, _ = gs.download_file_bytes(
            full_template_path,
            target_user_id=target_user_graph_id,
            drive_id=drive_id,
        )

        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = load_workbook(BytesIO(tpl_bytes))
        markers = []

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        text = str(cell.value).strip()
                        is_marker = False
                        
                        if text.endswith(':'):
                            is_marker = True
                        elif any(keyword in text.upper() for keyword in [
                            'DATOS', 'DETALLE', 'TOTAL', 'RESUMEN', 'CLIENTE',
                            'PAGOS', 'ITEMS', 'PRODUCTOS', 'SERVICIOS', 'FOOTER',
                            'HEADER', 'INFORMACIÓN'
                        ]):
                            is_marker = True
                        elif text.isupper() and len(text) > 3:
                            is_marker = True
                        
                        if is_marker:
                            col_letter = get_column_letter(cell.column)
                            markers.append({
                                "text": text,
                                "position": f"{col_letter}{cell.row}",
                                "sheet": sheet.title,
                                "row": cell.row,
                                "column": cell.column
                            })

        return jsonify({
            "message": "OK",
            "template_key": template.template_key,
            "markers_found": len(markers),
            "markers": markers,
            "correlation_id": corr_id
        }), 200

    except GraphAPIError as ge:
        status = ge.status_code if 400 <= (ge.status_code or 0) <= 599 else 502
        payload = {
            "error": ge.message,
            "correlation_id": corr_id,
        }
        if ge.ms_request_id:
            payload["ms_request_id"] = ge.ms_request_id
        return jsonify(payload), status

    except Exception as e:
        return jsonify({"error": str(e), "correlation_id": corr_id}), 500
