# excel_render.py
from io import BytesIO
from typing import Dict, Any
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _anchor_address_for(ws, addr: str) -> str:
    """Si addr cae dentro de un rango combinado, devuelve la celda ancla (min_row,min_col).
    Si no, devuelve addr tal cual.
    """
    row, col = coordinate_to_tuple(addr)  # (row:int, col:int)
    for mr in ws.merged_cells.ranges:
        # Checamos pertenencia con límites numéricos (sin usar "in")
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return f"{get_column_letter(mr.min_col)}{mr.min_row}"
    return addr

def fill_cells_in_memory(template_bytes: bytes, data: Dict[str, Any]) -> BytesIO:
    """
    Rellena celdas en un workbook a partir de un dict:
      - "A1": valor        → hoja ACTIVA del template
      - "Hoja1!B3": valor  → hoja 'Hoja1', celda B3
    Si la celda está combinada, escribe en la ancla del merge.
    """
    wb = load_workbook(BytesIO(template_bytes))
    for key, value in data.items():
        if "!" in key:
            sheet_name, addr = key.split("!", 1)
            ws = wb[sheet_name]  # KeyError si la hoja no existe
        else:
            ws = wb.active
            addr = key

        target_addr = _anchor_address_for(ws, addr)
        ws[target_addr].value = value

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out