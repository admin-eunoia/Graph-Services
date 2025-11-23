# excel_section_writer.py
"""
Sistema SIMPLIFICADO de escritura en Excel con secciones dinámicas.

3 funciones principales:
1. copiar_template() - Copia el Excel template
2. llenar_seccion() - Llena una sección con datos
3. guardar_excel() - Guarda el resultado
"""
from io import BytesIO
from typing import Dict, Any, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from copy import copy

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ==========================================
# FUNCIÓN 1: COPIAR TEMPLATE
# ==========================================

def copiar_template(template_bytes: bytes):
    """
    Copia el template Excel en memoria.
    
    Args:
        template_bytes: Bytes del archivo Excel original
    
    Returns:
        Workbook de openpyxl listo para editar
    
    Ejemplo:
        wb = copiar_template(template_bytes)
    """
    return load_workbook(BytesIO(template_bytes))


# ==========================================
# FUNCIÓN 2: LLENAR SECCIÓN
# ==========================================

def llenar_seccion(
    wb,
    marker: str,
    datos: Any,
    es_tabla: bool = False,
    columnas: Dict[str, int] = None
):
    """
    Llena una sección del Excel buscando un marcador.
    
    Args:
        wb: Workbook de openpyxl
        marker: Texto a buscar (ej: "Pagos:", "DATOS DEL CLIENTE:")
        datos: Dict para sección simple, List[Dict] para tabla
        es_tabla: True si es tabla (múltiples filas), False si es key-value simple
        columnas: Mapeo de campo -> offset de columna. Ej: {"fecha": 0, "monto": 1}
    
    Ejemplos:
        # Sección simple (key-value)
        llenar_seccion(
            wb,
            marker="DATOS DEL CLIENTE:",
            datos={"nombre": "ACME", "rfc": "ACM123"},
            es_tabla=False,
            columnas={"nombre": 0, "rfc": 1}
        )
        
        # Tabla (múltiples filas)
        llenar_seccion(
            wb,
            marker="Pagos:",
            datos=[
                {"fecha": "2025-01", "monto": 1000},
                {"fecha": "2025-02", "monto": 1500}
            ],
            es_tabla=True,
            columnas={"fecha": 0, "monto": 1}
        )
    """
    ws = wb.active
    columnas = columnas or {}
    
    # Desproteger temporalmente la hoja si está protegida
    estaba_protegida = ws.protection.sheet
    if estaba_protegida:
        ws.protection.sheet = False
        print(f"   ⚠️  Hoja temporalmente desprotegida para edición")
    
    # Buscar el marcador
    marker_row, marker_col = _buscar_marcador(ws, marker)
    if not marker_row:
        raise ValueError(f"No se encontró '{marker}' en el Excel")
    
    if es_tabla:
        # Llenar tabla (múltiples filas)
        _llenar_tabla(ws, marker_row, marker_col, datos, columnas)
    else:
        # Llenar valores simples (una sola fila)
        _llenar_valores(ws, marker_row, marker_col, datos, columnas)
    
    # Reproteger la hoja si estaba protegida
    if estaba_protegida:
        ws.protection.sheet = True
        print(f"   🔒 Hoja protegida nuevamente")


# ==========================================
# FUNCIÓN 3: GUARDAR EXCEL
# ==========================================

def guardar_excel(wb) -> BytesIO:
    """
    Guarda el workbook en memoria como BytesIO.
    
    Args:
        wb: Workbook de openpyxl
    
    Returns:
        BytesIO con el Excel completo
    
    Ejemplo:
        output = guardar_excel(wb)
        # Ahora puedes subirlo o guardarlo
    """
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==========================================
# FUNCIONES AUXILIARES (PRIVADAS)
# ==========================================

def _buscar_marcador(ws: Worksheet, marker: str) -> Tuple[Optional[int], Optional[int]]:
    """Busca el marcador en el Excel y retorna (fila, columna)."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and marker in str(cell.value):
                print(f"   ✓ Marcador '{marker}' encontrado en fila {cell.row}, columna {cell.column}")
                return (cell.row, cell.column)
    return (None, None)


def _escribir_en_celda(ws: Worksheet, fila: int, col: int, valor: Any):
    """Escribe en una celda, descombinándola si es necesario."""
    celda = ws.cell(fila, col)
    
    # Si es una celda combinada, NO descombinar - solo escribir en la celda principal
    if isinstance(celda, MergedCell):
        # Para celdas combinadas, encontrar la celda principal (top-left del rango)
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= fila <= merged_range.max_row and 
                merged_range.min_col <= col <= merged_range.max_col):
                # Obtener la celda principal del rango combinado
                celda_principal = ws.cell(merged_range.min_row, merged_range.min_col)
                print(f"   ℹ️  Escribiendo en celda combinada principal ({merged_range.min_row}, {merged_range.min_col}) para ({fila}, {col})")
                
                # Escribir el valor en la celda principal
                if isinstance(valor, str) and valor.startswith("="):
                    celda_principal.value = valor
                else:
                    celda_principal.value = valor
                return
    
    # Si no es celda combinada, escribir directamente
    if isinstance(valor, str) and valor.startswith("="):
        celda.value = valor
    else:
        celda.value = valor


def _llenar_valores(ws: Worksheet, marker_row: int, marker_col: int, datos: Dict, columnas: Dict):
    """Llena valores simples (key-value) en la fila siguiente al marcador."""
    fila_destino = marker_row + 1  # Siguiente fila después del marcador
    
    for campo, valor in datos.items():
        if campo not in columnas:
            continue
        
        col_offset = columnas[campo]
        col_destino = marker_col + col_offset
        
        # Escribir en la celda (descombinándola si es necesario)
        _escribir_en_celda(ws, fila_destino, col_destino, valor)


def _llenar_tabla(ws: Worksheet, marker_row: int, marker_col: int, filas: List[Dict], columnas: Dict):
    """Llena tabla (múltiples filas) después del marcador."""
    fila_inicio = marker_row + 2  # Saltar marcador + header
    
    # Insertar filas adicionales si es necesario
    num_filas_necesarias = len(filas)
    if num_filas_necesarias > 1:
        ws.insert_rows(fila_inicio + 1, num_filas_necesarias - 1)
    
    # Llenar cada fila
    for idx, fila_datos in enumerate(filas):
        fila_actual = fila_inicio + idx
        
        # Copiar formato de la fila template
        if idx > 0:
            _copiar_formato_fila(ws, fila_inicio, fila_actual)
        
        # Escribir datos
        for campo, valor in fila_datos.items():
            if campo not in columnas:
                continue
            
            col_offset = columnas[campo]
            col_destino = marker_col + col_offset
            
            # Escribir en la celda (descombinándola si es necesario)
            _escribir_en_celda(ws, fila_actual, col_destino, valor)


def _copiar_formato_fila(ws: Worksheet, fila_origen: int, fila_destino: int):
    """Copia el formato de una fila a otra."""
    for col in range(1, ws.max_column + 1):
        celda_origen = ws.cell(fila_origen, col)
        celda_destino = ws.cell(fila_destino, col)
        
        if celda_origen.has_style:
            celda_destino.font = copy(celda_origen.font)
            celda_destino.border = copy(celda_origen.border)
            celda_destino.fill = copy(celda_origen.fill)
            celda_destino.alignment = copy(celda_origen.alignment)


# ==========================================
# FUNCIÓN TODO-EN-UNO (OPCIONAL)
# ==========================================

def procesar_excel_completo(
    template_bytes: bytes,
    secciones: Dict[str, Any],
    configuracion: Dict[str, Dict]
) -> BytesIO:
    """
    Función todo-en-uno que procesa todas las secciones de un Excel.
    
    Args:
        template_bytes: Bytes del template
        secciones: Datos a escribir por sección
        configuracion: Config de cada sección (marker, columnas, es_tabla)
    
    Returns:
        BytesIO con el Excel procesado
    
    Ejemplo:
        output = procesar_excel_completo(
            template_bytes,
            secciones={
                "cliente": {"nombre": "ACME", "rfc": "ACM123"},
                "pagos": [
                    {"fecha": "2025-01", "monto": 1000},
                    {"fecha": "2025-02", "monto": 1500}
                ]
            },
            configuracion={
                "cliente": {
                    "marker": "DATOS DEL CLIENTE:",
                    "es_tabla": False,
                    "columnas": {"nombre": 0, "rfc": 1}
                },
                "pagos": {
                    "marker": "Pagos:",
                    "es_tabla": True,
                    "columnas": {"fecha": 0, "monto": 1}
                }
            }
        )
    """
    # 1. Copiar template
    wb = copiar_template(template_bytes)
    
    # 2. Llenar cada sección
    for nombre_seccion, datos in secciones.items():
        if nombre_seccion not in configuracion:
            continue
        
        config = configuracion[nombre_seccion]
        llenar_seccion(
            wb,
            marker=config["marker"],
            datos=datos,
            es_tabla=config.get("es_tabla", False),
            columnas=config.get("columnas", {})
        )
    
    # 3. Guardar
    return guardar_excel(wb)
