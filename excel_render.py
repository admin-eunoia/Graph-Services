# excel_render.py
from io import BytesIO
from typing import Dict, Any
from openpyxl import load_workbook

# MIME oficial de .xlsx (lo reutiliza Graph al subir)
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def fill_cells_in_memory(template_bytes: bytes, data: Dict[str, Any]) -> BytesIO:
    """
    Rellena celdas en un workbook a partir de un dict:
      - "A1": valor        → hoja ACTIVA del template
      - "Hoja1!B3": valor  → hoja 'Hoja1', celda B3
    Devuelve BytesIO con el .xlsx resultante (en memoria).
    """
    wb = load_workbook(BytesIO(template_bytes))
    for key, value in data.items():
        if "!" in key:
            sheet_name, addr = key.split("!", 1)
            ws = wb[sheet_name]  # KeyError si la hoja no existe
        else:
            ws = wb.active
            addr = key
        ws[addr].value = value

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
